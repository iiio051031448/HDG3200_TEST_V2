from openpyxl import Workbook
import os

EXPORT_XL_DIR_PATH="日志报表"


def file_exist(xl_path):
    return os.path.exists(xl_path)


class ExportXL:
    def __init__(self):
        print("-")
        self.workbook = Workbook()
        self.worksheet = None
        self.worksheet_nex_row = 1

    def create_header(self):
        # create header
        # mac operator start_time end_time test_id is_repeat result failed_info note
        sheet_header = ["mac", "操作者", "开始时间", "结束时间", "测试ID", "是否重复测试", "结果", "失败信息", "备注"]
        self.worksheet_nex_row = 1
        for j, h in enumerate(sheet_header):
            self.worksheet.cell(row=1, column=j + 1, value=h)
            self.worksheet_nex_row = 2

    def active_now(self, new_sheet_name=None):
        self.worksheet = self.workbook.active
        if new_sheet_name:
            self.worksheet.title = new_sheet_name
        print(self.worksheet)
        self.create_header()

    def active_new(self, sheet_name):
        self.worksheet = self.workbook.create_sheet(sheet_name)
        self.create_header()

    def save(self, xl_path):
        if not os.path.exists(EXPORT_XL_DIR_PATH):
            os.mkdir(EXPORT_XL_DIR_PATH)

        self.workbook.save(xl_path)

    def add_row(self, row_data):
        # row_data = [mac, operator, start_time, end_time, test_id, is_repeat, result, failed_info, note]
        if not self.worksheet:
            print("worksheet is not active !")
            return
        for j, h in enumerate(row_data):
            self.worksheet.cell(row=self.worksheet_nex_row, column=j + 1, value=h)
        self.worksheet_nex_row += 1

    def add_item(self, mac, operator, start_time, end_time, test_id, is_repeat, result, failed_info, note):
        row_data = [mac, operator, start_time, end_time, test_id, is_repeat, result, failed_info, note]
        self.add_row(row_data)

if __name__ == "__main__":
    exl = ExportXL()
    exl.active_now("成功")
    exl.add_item("mac", "operator", "start_time", "end_time", "test_id", "is_repeat", "result", "failed_info", "note")

    exl.active_new("失败")
    exl.add_item("mac", "operator", "start_time", "end_time", "test_id", "is_repeat", "result", "failed_info", "note")

    exl.save('./' + EXPORT_XL_DIR_PATH + '/' + 'H2222.xlsx')

